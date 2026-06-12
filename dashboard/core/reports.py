"""
Génération de rapports PDF et Excel — SIEM Africa Dashboard.

Produit un rapport de sécurité (hebdomadaire ou personnalisé) à partir des
données réelles de la base. Deux formats : PDF (reportlab) et Excel (openpyxl).
Les fichiers sont écrits dans un dossier de rapports et tracés dans la table
reports du Module 2.
"""
import os
import uuid
from datetime import datetime, timedelta

from django.db import connection
from django.conf import settings as dj_settings

from . import stats

REPORTS_DIR = os.environ.get(
    "SIEM_REPORTS_PATH",
    os.path.join(dj_settings.BASE_DIR, "reports_output"),
)


def _ensure_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _collect_data(days=7):
    """Rassemble les données du rapport sur la période donnée."""
    hours = days * 24
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "period_days": days,
        "severity": stats.severity_counts(hours),
        "total_alerts": stats.alerts_total(hours),
        "incidents_open": stats.incidents_open_count(),
        "blocked_ips": stats.blocked_ips_count(active_only=False),
        "score": stats.security_score(),
        "top_categories": stats.top_categories(hours, 10),
        "top_ips": stats.top_attacking_ips(hours, 10),
    }


# --- PDF ---------------------------------------------------------------------
def generate_pdf(days=7, lang="fr"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    _ensure_dir()
    data = _collect_data(days)
    fname = f"rapport_securite_{datetime.now():%Y%m%d_%H%M%S}.pdf"
    path = os.path.join(REPORTS_DIR, fname)

    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#1F4E79"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E79"))
    elems = []

    T = {
        "title": "Rapport de sécurité SIEM Africa" if lang == "fr" else "SIEM Africa Security Report",
        "period": f"Période : {days} derniers jours" if lang == "fr" else f"Period: last {days} days",
        "generated": f"Généré le {data['generated_at']}" if lang == "fr" else f"Generated on {data['generated_at']}",
        "summary": "Synthèse" if lang == "fr" else "Summary",
        "score": "Score de sécurité" if lang == "fr" else "Security score",
        "total": "Total alertes" if lang == "fr" else "Total alerts",
        "incidents": "Incidents ouverts" if lang == "fr" else "Open incidents",
        "blocked": "IPs bloquées" if lang == "fr" else "Blocked IPs",
        "by_sev": "Répartition par sévérité" if lang == "fr" else "By severity",
        "top_cat": "Top catégories d'attaques" if lang == "fr" else "Top attack categories",
        "top_ips": "Top IPs attaquantes" if lang == "fr" else "Top attacking IPs",
    }

    elems.append(Paragraph(T["title"], title_style))
    elems.append(Paragraph(T["period"], styles["Normal"]))
    elems.append(Paragraph(T["generated"], styles["Normal"]))
    elems.append(Spacer(1, 0.6*cm))

    # Synthèse
    elems.append(Paragraph(T["summary"], h2))
    summary = [
        [T["score"], f"{data['score']['score']}/100"],
        [T["total"], str(data["total_alerts"])],
        [T["incidents"], str(data["incidents_open"])],
        [T["blocked"], str(data["blocked_ips"])],
    ]
    tbl = Table(summary, colWidths=[8*cm, 6*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F0F4F8")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("FONTSIZE", (0,0), (-1,-1), 11),
        ("PADDING", (0,0), (-1,-1), 8),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 0.5*cm))

    # Sévérités
    elems.append(Paragraph(T["by_sev"], h2))
    sev = data["severity"]
    sev_tbl = Table(
        [["CRITICAL","HIGH","MEDIUM","LOW","INFO"],
         [sev["CRITICAL"],sev["HIGH"],sev["MEDIUM"],sev["LOW"],sev["INFO"]]],
        colWidths=[2.8*cm]*5,
    )
    sev_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTSIZE", (0,0), (-1,-1), 11),
        ("PADDING", (0,0), (-1,-1), 8),
    ]))
    elems.append(sev_tbl)
    elems.append(Spacer(1, 0.5*cm))

    # Top catégories
    if data["top_categories"]:
        elems.append(Paragraph(T["top_cat"], h2))
        rows = [["Catégorie" if lang=="fr" else "Category", "Nombre" if lang=="fr" else "Count"]]
        rows += [[c["categorie"], str(c["n"])] for c in data["top_categories"]]
        ct = Table(rows, colWidths=[10*cm, 4*cm])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTSIZE", (0,0), (-1,-1), 10), ("PADDING", (0,0), (-1,-1), 6),
        ]))
        elems.append(ct)
        elems.append(Spacer(1, 0.5*cm))

    # Top IPs
    if data["top_ips"]:
        elems.append(Paragraph(T["top_ips"], h2))
        rows = [["IP", "Pays" if lang=="fr" else "Country", "Alertes" if lang=="fr" else "Alerts"]]
        rows += [[i["ip"], i.get("country_code") or "—", str(i["n"])] for i in data["top_ips"]]
        it = Table(rows, colWidths=[6*cm, 4*cm, 4*cm])
        it.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTSIZE", (0,0), (-1,-1), 10), ("PADDING", (0,0), (-1,-1), 6),
        ]))
        elems.append(it)

    doc.build(elems)
    _register_report("Rapport de sécurité", "pdf", path, days)
    return path, fname


# --- Excel -------------------------------------------------------------------
def generate_excel(days=7, lang="fr"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _ensure_dir()
    data = _collect_data(days)
    fname = f"rapport_securite_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = os.path.join(REPORTS_DIR, fname)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    # Feuille synthèse
    ws = wb.active
    ws.title = "Synthèse" if lang == "fr" else "Summary"
    ws["A1"] = "Rapport de sécurité SIEM Africa" if lang == "fr" else "SIEM Africa Security Report"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
    ws["A2"] = f"Généré le {data['generated_at']}"
    ws["A3"] = f"Période : {days} jours" if lang == "fr" else f"Period: {days} days"

    summary = [
        ("Score de sécurité", f"{data['score']['score']}/100"),
        ("Total alertes", data["total_alerts"]),
        ("Incidents ouverts", data["incidents_open"]),
        ("IPs bloquées", data["blocked_ips"]),
        ("CRITICAL", data["severity"]["CRITICAL"]),
        ("HIGH", data["severity"]["HIGH"]),
        ("MEDIUM", data["severity"]["MEDIUM"]),
        ("LOW", data["severity"]["LOW"]),
        ("INFO", data["severity"]["INFO"]),
    ]
    row = 5
    for label, val in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    # Feuille top catégories
    ws2 = wb.create_sheet("Catégories" if lang == "fr" else "Categories")
    ws2.append(["Catégorie", "Nombre"])
    for c in ws2[1]:
        c.fill = header_fill; c.font = header_font
    for cat in data["top_categories"]:
        ws2.append([cat["categorie"], cat["n"]])
    ws2.column_dimensions["A"].width = 30

    # Feuille top IPs
    ws3 = wb.create_sheet("IPs")
    ws3.append(["IP", "Pays", "Alertes"])
    for c in ws3[1]:
        c.fill = header_fill; c.font = header_font
    for ip in data["top_ips"]:
        ws3.append([ip["ip"], ip.get("country_code") or "—", ip["n"]])
    ws3.column_dimensions["A"].width = 18

    wb.save(path)
    _register_report("Rapport de sécurité", "excel", path, days)
    return path, fname


def _register_report(title, fmt, path, days):
    """Enregistre le rapport dans la table reports.

    La table M2 contraint report_type (DAILY/WEEKLY/MONTHLY/CUSTOM/INCIDENT) et
    format (PDF/CSV/HTML/JSON). Le format Excel n'étant pas prévu par la
    contrainte, on stocke un format autorisé dans la colonne et on conserve le
    format réel dans metadata (JSON). Le téléchargement se fonde sur l'extension
    réelle du fichier (file_path)."""
    import json as _json
    try:
        size = os.path.getsize(path)
        # report_type selon la période
        if days <= 1:
            rtype = "DAILY"
        elif days <= 7:
            rtype = "WEEKLY"
        elif days <= 31:
            rtype = "MONTHLY"
        else:
            rtype = "CUSTOM"
        # format colonne : valeur autorisée par la contrainte ; le vrai format
        # (pdf/excel) est dans metadata et déductible de l'extension du fichier
        col_format = "PDF" if fmt == "pdf" else "CSV"
        meta = _json.dumps({"real_format": fmt})
        with connection.cursor() as cur:
            cur.execute(
                """INSERT INTO reports (report_uuid, title, report_type, format,
                   period_start, period_end, file_path, file_size_bytes,
                   generation_status, metadata, created_at, completed_at)
                   VALUES (%s, %s, %s, %s, datetime('now', %s),
                           datetime('now'), %s, %s, 'COMPLETED', %s, datetime('now'), datetime('now'))""",
                [str(uuid.uuid4()), title, rtype, col_format, f"-{days} days",
                 path, size, meta],
            )
    except Exception:
        pass


def list_reports(limit=50):
    with connection.cursor() as cur:
        cur.execute(
            """SELECT id, title, report_type, format, file_size_bytes,
                      generation_status, created_at, file_path, metadata
               FROM reports ORDER BY created_at DESC LIMIT %s""",
            [limit],
        )
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    # Déduire le format réel depuis l'extension ou metadata
    import json as _json
    for r in rows:
        real = None
        if r.get("metadata"):
            try:
                real = _json.loads(r["metadata"]).get("real_format")
            except (ValueError, TypeError):
                real = None
        if not real and r.get("file_path"):
            real = "excel" if r["file_path"].endswith(".xlsx") else "pdf"
        r["real_format"] = real or "pdf"
    return rows


def get_report_path(report_id):
    with connection.cursor() as cur:
        cur.execute("SELECT file_path, format FROM reports WHERE id=%s", [report_id])
        row = cur.fetchone()
        return (row[0], row[1]) if row else (None, None)
